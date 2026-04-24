"""
req_report.py

Generates a requirements list and hierarchy diagrams from a SysML v2 model.

Outputs:
  1. requirements.md  — flat list with IDs as enumerators, doc annotations
                        rendered with labels; unnamed doc is the requirement
                        text, 'doc Rationale' is the rationale block.
  2. req_hierarchy_<ID>.png — one Graphviz LR diagram per top-level requirement,
     showing the full derivation tree with the current node highlighted.

Usage:
    python __Tools/req_report.py <model_dir>
    python __Tools/req_report.py <model_dir> --format xlsx
    python __Tools/req_report.py <model_dir> --output ./reports

Dependencies:
    pip install graphviz openpyxl   (openpyxl only needed for --format xlsx)
    graphviz system package must also be installed:
      Linux:  sudo apt install graphviz
      macOS:  brew install graphviz
      Windows: https://graphviz.org/download/

SysIDE API notes (SysIDE 0.8.x / SysML v2 2025-07):
  - RequirementUsage.req_id         -> the <'A.1'> short-name identifier string
  - RequirementUsage.nested_requirements -> subrequirements (from Usage)
  - element.documentation           -> list of Documentation elements
  - doc.declared_name               -> the name after 'doc', or None for unnamed
  - doc.body                        -> the /* ... */ text content
  - top_elements_from(path) scopes to user model files only

Doc Convention (SRS_Definitions.sysml):
  Every requirement usage must carry:
    doc           /* <normative "shall" text> */
    doc Rationale /* <rationale prose> */
  Plus a short-name ID:  requirement <'REQ-001'> myReq : SomeReqDef { ... }
  These are validated by req_validate.py; this tool reports them as-found.
"""

import sys
import re
import argparse
import textwrap
from pathlib import Path
from dataclasses import dataclass, field
import syside
from syside.preview import open_model


# ── Data model ────────────────────────────────────────────────────────────────

@dataclass
class DocEntry:
    """A single doc annotation: a name (or None) and its body text."""
    name: str | None   # None for unnamed doc; 'Rationale' for doc Rationale
    body: str          # stripped body text, leading /* */ removed


@dataclass
class ReqNode:
    """A single requirement with its ID, name, doc annotations, and children."""
    req_id:   str               # e.g. "A", "A.1", "A.2.1"
    name:     str               # declared_name, e.g. "requirementA"
    docs:     list[DocEntry]    # all doc annotations in declaration order
    children: list["ReqNode"] = field(default_factory=list)

    @property
    def req_text(self) -> str:
        """The unnamed doc body — the normative requirement statement."""
        for d in self.docs:
            if d.name is None:
                return d.body
        return ""

    @property
    def rationale(self) -> str:
        """The 'doc Rationale' body."""
        for d in self.docs:
            if d.name and d.name.lower() == "rationale":
                return d.body
        return ""

    @property
    def label(self) -> str:
        """Short label for display: ID if present, else name."""
        return self.req_id if self.req_id else self.name

    @property
    def node_id(self) -> str:
        """Graphviz-safe node identifier (no spaces or special chars)."""
        safe = re.sub(r"[^A-Za-z0-9_]", "_", self.req_id or self.name)
        return f"n_{safe}"


# ── Helpers ───────────────────────────────────────────────────────────────────

def get_name(element) -> str:
    if element is None:
        return ""
    name = element.declared_name
    if name:
        return name
    qn = element.qualified_name
    return str(qn) if qn else ""


def get_req_id(req: syside.RequirementUsage) -> str:
    """
    Return the requirement ID from req_id or short_name.
    The <'A.1'> syntax sets short_name; req_id is the spec accessor for it.
    Falls back to declared_name if neither is set.
    """
    try:
        rid = req.req_id
        if rid:
            return str(rid).strip("'\"")
    except Exception:
        pass
    try:
        sn = req.short_name
        if sn:
            return str(sn).strip("'\"")
    except Exception:
        pass
    return get_name(req)


def _strip_doc_markers(body: str) -> str:
    """Remove /* */ comment markers and normalize leading * on continuation lines."""
    text = str(body).strip()
    # Strip opening /* with any number of asterisks
    text = re.sub(r"^/\*+\s*", "", text)
    # Strip closing */ with any number of asterisks
    text = re.sub(r"\s*\*+/$", "", text)
    # Remove leading ' * ' on interior lines (JavaDoc / SysML style)
    lines = text.splitlines()
    cleaned = []
    for line in lines:
        cleaned.append(re.sub(r"^\s*\*\s?", "", line))
    return "\n".join(cleaned).strip()


def get_all_docs(req: syside.RequirementUsage) -> list[DocEntry]:
    """
    Return all doc annotations on this requirement, in declaration order.
    Each DocEntry carries:
      name — the declared name after 'doc' (e.g. 'Rationale'), or None for unnamed
      body — the cleaned text content
    """
    entries: list[DocEntry] = []
    try:
        for doc in req.documentation.collect():
            body = doc.body
            if not body:
                continue
            cleaned = _strip_doc_markers(body)
            if not cleaned:
                continue
            doc_name = getattr(doc, "declared_name", None) or None
            entries.append(DocEntry(name=doc_name, body=cleaned))
    except Exception:
        pass
    return entries


def collect_typed(root, std_type, results: list):
    """Depth-first collection of all elements matching std_type."""
    if root.isinstance(std_type):
        results.append(root.cast(std_type))
    if root.isinstance(syside.Namespace.STD):
        ns = root.cast(syside.Namespace.STD)
        for member in ns.owned_members.collect():
            collect_typed(member, std_type, results)


def is_plain_req(r) -> bool:
    """Filter to plain RequirementUsage — exclude satisfy, concern, viewpoint."""
    return (
        not r.isinstance(syside.SatisfyRequirementUsage.STD)
        and not r.isinstance(syside.ConcernUsage.STD)
    )


def build_req_tree(req: syside.RequirementUsage) -> ReqNode:
    """Recursively build a ReqNode tree from a RequirementUsage."""
    node = ReqNode(
        req_id=get_req_id(req),
        name=get_name(req),
        docs=get_all_docs(req),
    )
    try:
        for child in req.nested_requirements.collect():
            if not child.isinstance(syside.RequirementUsage.STD):
                continue
            child_req = child.cast(syside.RequirementUsage.STD)
            if is_plain_req(child_req):
                node.children.append(build_req_tree(child_req))
    except Exception:
        pass
    return node


def flatten(node: ReqNode, depth: int = 0) -> list[tuple[int, ReqNode]]:
    """Return (depth, node) pairs in pre-order."""
    result = [(depth, node)]
    for child in node.children:
        result.extend(flatten(child, depth + 1))
    return result


# ── Markdown output ───────────────────────────────────────────────────────────

def _md_doc_block(docs: list[DocEntry]) -> str:
    """
    Render doc annotations for the Markdown list.
    Unnamed doc → no label prefix (it IS the requirement text).
    Named doc   → bold label prefix, e.g. **Rationale:** ...
    All bodies are single-line for list output; paragraph breaks become spaces.
    """
    parts = []
    for d in docs:
        # Collapse multiline to single line for list output
        single_line = " ".join(d.body.split())
        if d.name is None:
            parts.append(single_line)
        else:
            parts.append(f"**{d.name}:** {single_line}")
    return "\n\n  ".join(parts)


def write_markdown(roots: list[ReqNode], output_dir: Path) -> Path:
    out = output_dir / "requirements.md"
    lines = ["# Requirements\n"]
    for root in roots:
        for depth, node in flatten(root):
            indent = "  " * depth
            header_level = min(depth + 2, 6)
            hashes = "#" * header_level
            label = node.label
            name_part = f" `{node.name}`" if node.name and node.name != label else ""
            lines.append(f"{indent}{hashes} {label}{name_part}\n")
            if node.docs:
                doc_text = _md_doc_block(node.docs)
                for part_line in doc_text.splitlines():
                    lines.append(f"{indent}{part_line}\n")
            else:
                lines.append(f"{indent}_No documentation provided._\n")
            lines.append("")
    out.write_text("".join(lines), encoding="utf-8")
    return out


# ── Excel output ──────────────────────────────────────────────────────────────

def write_xlsx(roots: list[ReqNode], output_dir: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1F4E79")
    wrap_align   = Alignment(wrap_text=True, vertical="top")

    headers = ["ID", "Name", "Depth", "Requirement Text", "Rationale", "Additional Docs"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = wrap_align

    row_num = 2
    for root in roots:
        for depth, node in flatten(root):
            # Collect additional (non-text, non-rationale) docs
            extra_docs = [
                f"{d.name}: {' '.join(d.body.split())}"
                for d in node.docs
                if d.name is not None and d.name.lower() != "rationale"
            ]
            ws.cell(row=row_num, column=1, value=node.req_id or "").alignment = wrap_align
            ws.cell(row=row_num, column=2, value=node.name or "").alignment = wrap_align
            ws.cell(row=row_num, column=3, value=depth).alignment = wrap_align
            ws.cell(row=row_num, column=4, value=node.req_text or "").alignment = wrap_align
            ws.cell(row=row_num, column=5, value=node.rationale or "").alignment = wrap_align
            ws.cell(row=row_num, column=6, value="\n".join(extra_docs)).alignment = wrap_align
            row_num += 1

    col_widths = [18, 30, 8, 70, 70, 50]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    out = output_dir / "requirements.xlsx"
    wb.save(out)
    return out


# ── Graphviz diagrams ─────────────────────────────────────────────────────────

def _graphviz_label(node: ReqNode, highlight: bool) -> str:
    """Build an HTML-like Graphviz label: ID on top, short doc preview below."""
    req_text = node.req_text
    if req_text:
        preview = textwrap.shorten(req_text, width=48, placeholder="…")
        preview_escaped = (
            preview.replace("&", "&amp;")
                   .replace("<", "&lt;")
                   .replace(">", "&gt;")
        )
        body = (
            f'<TABLE BORDER="0" CELLBORDER="0" CELLSPACING="2">'
            f'<TR><TD><B>{node.label}</B></TD></TR>'
            f'<TR><TD><FONT POINT-SIZE="9">{preview_escaped}</FONT></TD></TR>'
            f'</TABLE>'
        )
    else:
        body = f'<B>{node.label}</B>'
    return f"<{body}>"


def write_diagram(root: ReqNode, highlight: ReqNode, output_dir: Path) -> Path:
    """
    Render one hierarchy diagram for `root`, highlighting the `highlight` node.

    Visibility rule (per ReqNode's ancestry):
      - All ancestors of `highlight` up to `root`
      - All siblings at every ancestor level
      - Direct children of `highlight` only (no grandchildren)
    """
    try:
        import graphviz
    except ImportError:
        print("graphviz Python package not installed. Run: pip install graphviz",
              file=sys.stderr)
        return None

    # Collect which nodes are visible
    # Build parent map for the whole subtree under root
    parent_map: dict[str, ReqNode] = {}
    def _index(n: ReqNode, parent: ReqNode | None = None):
        if parent:
            parent_map[n.node_id] = parent
        for c in n.children:
            _index(c, n)
    _index(root)

    # Ancestors of highlight (exclusive of highlight itself)
    def ancestors(n: ReqNode) -> list[ReqNode]:
        chain = []
        cur = parent_map.get(n.node_id)
        while cur is not None:
            chain.append(cur)
            cur = parent_map.get(cur.node_id)
        return chain

    ancestor_ids = {a.node_id for a in ancestors(highlight)}
    ancestor_ids.add(root.node_id)

    # Siblings at every ancestor level (all children of ancestors)
    sibling_ids: set[str] = set()
    for anc in [root] + [n for n in flatten(root) if n[1].node_id in ancestor_ids]:
        for c in anc[1].children:
            sibling_ids.add(c.node_id)

    # Direct children of highlight
    child_ids = {c.node_id for c in highlight.children}

    visible_ids = ancestor_ids | sibling_ids | child_ids | {highlight.node_id}

    dot = graphviz.Digraph(
        graph_attr={
            "rankdir": "LR",
            "splines": "spline",
            "bgcolor": "white",
            "fontname": "Helvetica",
        },
        node_attr={
            "shape": "box",
            "style": "filled",
            "fontname": "Helvetica",
            "fontsize": "10",
        },
        edge_attr={"color": "#555555"},
    )

    def add_nodes_edges(n: ReqNode, parent_id: str | None = None):
        if n.node_id not in visible_ids:
            return
        is_highlight = n.node_id == highlight.node_id
        dot.node(
            n.node_id,
            label=_graphviz_label(n, is_highlight),
            fillcolor="#1F4E79" if is_highlight else "#DDEEFF",
            fontcolor="white" if is_highlight else "black",
        )
        if parent_id:
            dot.edge(parent_id, n.node_id)
        for child in n.children:
            add_nodes_edges(child, n.node_id)

    add_nodes_edges(root)

    safe_id = re.sub(r"[^A-Za-z0-9_\-]", "_", highlight.label)
    out_path = output_dir / f"req_hierarchy_{safe_id}"
    dot.render(str(out_path), format="png", cleanup=True)
    return Path(str(out_path) + ".png")


# ── Main ──────────────────────────────────────────────────────────────────────

def run(model_dir: Path, fmt: str, output_dir: Path):
    print(f"Opening model at: {model_dir}")
    with open_model(str(model_dir)) as model:
        all_reqs: list[syside.RequirementUsage] = []
        for element in model.top_elements_from(str(model_dir)):
            collect_typed(element, syside.RequirementUsage.STD, all_reqs)

        plain_reqs = [r for r in all_reqs if is_plain_req(r)]
        print(f"Found {len(plain_reqs)} requirement usage(s).")

        # Separate top-level from nested: a top-level req has no parent req in the list
        all_req_ids_set = {id(r) for r in plain_reqs}
        nested_ids: set[int] = set()
        for r in plain_reqs:
            try:
                for child in r.nested_requirements.collect():
                    if child.isinstance(syside.RequirementUsage.STD):
                        nested_ids.add(id(child.cast(syside.RequirementUsage.STD)))
            except Exception:
                pass
        top_level = [r for r in plain_reqs if id(r) not in nested_ids]

        roots = [build_req_tree(r) for r in top_level]

        output_dir.mkdir(parents=True, exist_ok=True)

        if fmt == "xlsx":
            out = write_xlsx(roots, output_dir)
        else:
            out = write_markdown(roots, output_dir)
        print(f"  Written: {out}")

        # Diagrams — one per node in every tree
        for root in roots:
            for _, node in flatten(root):
                png = write_diagram(root, node, output_dir)
                if png:
                    print(f"  Diagram: {png.name}")


def main():
    parser = argparse.ArgumentParser(
        description="Generate SysML v2 requirements list and hierarchy diagrams."
    )
    parser.add_argument(
        "model_dir",
        help="Path to the SysML v2 model directory.",
    )
    parser.add_argument(
        "--format", "-f",
        choices=["md", "xlsx"],
        default="md",
        help="Output format: md (default) or xlsx.",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output directory (default: __output in current working directory).",
    )
    args = parser.parse_args()

    model_dir = Path(args.model_dir).resolve()
    if not model_dir.is_dir():
        print(f"Error: model directory not found: {model_dir}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output).resolve() if args.output else Path.cwd() / "__output"
    output_dir.mkdir(parents=True, exist_ok=True)

    run(model_dir, args.format, output_dir)


if __name__ == "__main__":
    main()
